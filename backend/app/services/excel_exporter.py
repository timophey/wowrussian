"""
Excel export service for analysis results.
"""
from typing import Dict, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import io
import logging
import tempfile
import os

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Exports analysis results to Excel format."""
    
    @staticmethod
    def export_analysis(
        analysis_data: dict,
        selected_statuses: Optional[List[str]] = None,
        page_url: Optional[str] = None,
        language: str = "ru"
    ) -> bytes:
        """
        Export analysis results to XLSX format.
        
        Args:
            analysis_data: Analysis results data (same structure as fz168 response)
            selected_statuses: List of statuses to include. If None/empty, all statuses are included
            page_url: URL of the analyzed page (will be shown in first column of each row)
            language: Language code for headers (ru or en)
            
        Returns:
            Bytes containing the XLSX file
        """
        try:
            all_words = analysis_data.get("all_words", [])
            total_words = len(all_words)
            logger.info(f"ExcelExporter: Starting export with {total_words} words, language={language}")

            # Filter words by selected statuses if provided (empty array means no words)
            if selected_statuses is not None and len(selected_statuses) > 0:
                filtered_words = [w for w in all_words if w.get("status") in selected_statuses]
            else:
                filtered_words = all_words
            
            # Get page URL from analysis_data if not explicitly provided
            if not page_url:
                source_info = analysis_data.get("source_info", {})
                if source_info and source_info.get("url"):
                    page_url = source_info["url"]
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Analysis Results"

            total_rows = len(filtered_words)
            # For large datasets, use minimal styling to speed up save
            use_styling = total_rows <= 10000

            if use_styling:
                # Define styles only for smaller datasets
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")

                status_styles = {
                    "ok": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"), "font_color": "006100"},
                    "prohibited": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"), "font_color": "9C0006"},
                    "foreign": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"), "font_color": "9C6500"},
                    "normative_violation": {"fill": PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid"), "font_color": "663300"},
                    "unknown": {"fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"), "font_color": "4D4D4D"},
                    "foreign_with_alternative": {"fill": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"), "font_color": "7F6000"},
                }

                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            else:
                logger.info(f"ExcelExporter: Using minimal styling for large dataset ({total_rows} rows)")
                # Use simple default styles for large datasets
                header_font = Font(bold=True)
                header_fill = None
                header_alignment = Alignment(horizontal="center")
                status_styles = {}
                thin_border = None
            
            # Localized headers based on language
            headers_map = {
                "ru": ["Ссылка на страницу", "Слово", "Кол-во", "Статус", "Словари", "Категория", "Рекомендация / Статья"],
                "en": ["Page URL", "Word", "Count", "Status", "Dictionaries", "Category", "Recommendation / Article"]
            }
            headers = headers_map.get(language, headers_map["ru"])
            
            # Write headers (row 1)
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                if header_fill:
                    cell.fill = header_fill
                cell.alignment = header_alignment
                if thin_border:
                    cell.border = thin_border
            
            # Write data rows (starting from row 2)
            current_row = 2
            total_rows = len(filtered_words)
            log_interval = max(10000, total_rows // 20)  # Log every 10k rows or 5% of data
            use_styling = total_rows <= 10000

            for idx, word_data in enumerate(filtered_words):
                # Log progress periodically for large datasets
                if idx % log_interval == 0:
                    logger.info(f"Excel export: processing row {idx + 1} of {total_rows} ({((idx + 1) / total_rows * 100):.1f}%)")

                word = word_data.get("word", "")
                count = word_data.get("count", 1)
                status = word_data.get("status", "")
                
                # Process dictionaries - strip user_ prefix and get unique values
                raw_dictionaries = word_data.get("dictionaries", [])
                if raw_dictionaries and isinstance(raw_dictionaries, list):
                    processed_dicts = []
                    for d in raw_dictionaries:
                        if isinstance(d, str):
                            d = d.removeprefix("user_")
                        if d not in processed_dicts:
                            processed_dicts.append(d)
                    dictionaries_str = ", ".join(processed_dicts)
                else:
                    dictionaries_str = ""
                
                categories = word_data.get("categories", [])
                category_str = ", ".join(categories) if categories else ""

                # Get recommendation or law article
                recommendation = ""
                if word_data.get("law_article"):
                    recommendation = f"Статья: {word_data['law_article']}"
                elif word_data.get("recommendation"):
                    recommendation = word_data["recommendation"]
                elif status == "ok":
                    recommendation = "Соответствует нормам" if language == "ru" else "Complies with standards"

                # Column 1: Page URL (as plain text for large datasets to avoid hyperlink overhead)
                word_page_url = word_data.get("page_url")
                final_url = word_page_url or page_url
                ws.cell(row=current_row, column=1, value=final_url or "")

                # Column 2: Word
                ws.cell(row=current_row, column=2, value=word)

                # Column 3: Count
                ws.cell(row=current_row, column=3, value=count)

                # Column 4: Status
                ws.cell(row=current_row, column=4, value=status)

                # Column 5: Dictionaries
                ws.cell(row=current_row, column=5, value=dictionaries_str)

                # Column 6: Category
                ws.cell(row=current_row, column=6, value=category_str)

                # Column 7: Recommendation/Article
                ws.cell(row=current_row, column=7, value=recommendation)

                # Apply styling only for smaller datasets
                if use_styling and thin_border:
                    for col in range(1, 8):
                        ws.cell(row=current_row, column=col).border = thin_border

                    if status in status_styles:
                        status_cell = ws.cell(row=current_row, column=4)
                        status_cell.fill = status_styles[status]["fill"]
                        status_cell.font = Font(color=status_styles[status]["font_color"], bold=True)

                current_row += 1
            
            logger.info(f"Excel export: completed writing {total_rows} rows")
            
            # Optimization: Skip auto-column width adjustment for large datasets (>10k rows)
            # This is the slowest operation in openpyxl for large files
            total_rows = current_row - 1
            if total_rows <= 10000:
                logger.info(f"ExcelExporter: Starting column width adjustment for {total_rows} rows")
                column_widths = {
                    1: 40,  # Page URL
                    2: 30,  # Word
                    3: 12,  # Count
                    4: 20,  # Status
                    5: 30,  # Dictionaries
                    6: 25,  # Category
                    7: 50   # Recommendation/Article
                }
                for col_idx, width in column_widths.items():
                    ws.column_dimensions[get_column_letter(col_idx)].width = width
                logger.info(f"ExcelExporter: Column widths set")
            else:
                logger.info(f"ExcelExporter: Skipping column width adjustment for large dataset ({total_rows} rows)")

            # Freeze the header row (row 1) - this is fast
            ws.freeze_panes = "A2"
            logger.info(f"ExcelExporter: Frozen panes")

            # Add auto-filter only for moderate sized datasets
            if total_rows <= 50000:
                if total_rows > 0:  # At least one data row
                    last_row = current_row - 1
                    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"
                else:
                    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
                logger.info(f"ExcelExporter: Set auto-filter")
            else:
                logger.info(f"ExcelExporter: Skipping auto-filter for large dataset ({total_rows} rows)")

            # Create summary sheet with aggregated statistics per page
            logger.info(f"ExcelExporter: Creating summary sheet")
            summary_ws = wb.create_sheet(title="Summary" if language == "en" else "Сводка")
            
            # Define summary headers based on language
            summary_headers_map = {
                "ru": [
                    "Ссылка на страницу",
                    "Всего слов (сумма)",
                    "Уникальных слов",
                    "Иностранные слова",
                    "Запрещенные слова",
                    "Нарушения нормативов",
                    "Соответствующие слова"
                ],
                "en": [
                    "Page URL",
                    "Total Words (sum)",
                    "Unique Words",
                    "Foreign Words",
                    "Prohibited Words",
                    "Normative Violations",
                    "Compliant Words"
                ]
            }
            summary_headers = summary_headers_map.get(language, summary_headers_map["ru"])
            
            # Apply styling to summary headers
            summary_header_font = Font(bold=True, color="FFFFFF")
            summary_header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            summary_header_alignment = Alignment(horizontal="center", vertical="center")
            summary_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Write summary headers
            for col_idx, header in enumerate(summary_headers, 1):
                cell = summary_ws.cell(row=1, column=col_idx, value=header)
                cell.font = summary_header_font
                cell.fill = summary_header_fill
                cell.alignment = summary_header_alignment
                cell.border = summary_border
            
            # Aggregate statistics by page_url
            page_stats = {}
            for word_data in filtered_words:
                page_url = word_data.get("page_url") or page_url or "Unknown"
                count = word_data.get("count", 1)
                status = word_data.get("status", "")
                word = word_data.get("word", "")
                
                if page_url not in page_stats:
                    page_stats[page_url] = {
                        'total_words': 0,
                        'unique_words': set(),
                        'foreign_count': 0,
                        'prohibited_count': 0,
                        'normative_count': 0,
                        'ok_count': 0,
                        'words': {}  # For tracking unique words with their counts
                    }
                
                stats = page_stats[page_url]
                stats['total_words'] += count
                stats['unique_words'].add(word)
                
                # Track word counts for unique word calculation
                if word not in stats['words']:
                    stats['words'][word] = {'count': count, 'status': status}
                else:
                    stats['words'][word]['count'] += count
                
                # Count by status
                if status == "foreign":
                    stats['foreign_count'] += count
                elif status == "prohibited":
                    stats['prohibited_count'] += count
                elif status == "normative_violation":
                    stats['normative_count'] += count
                elif status == "ok":
                    stats['ok_count'] += count
            
            # Write summary data rows
            summary_row = 2
            for page_url, stats in sorted(page_stats.items()):
                unique_words_count = len(stats['unique_words'])
                
                summary_ws.cell(row=summary_row, column=1, value=page_url)
                summary_ws.cell(row=summary_row, column=2, value=stats['total_words'])
                summary_ws.cell(row=summary_row, column=3, value=unique_words_count)
                summary_ws.cell(row=summary_row, column=4, value=stats['foreign_count'])
                summary_ws.cell(row=summary_row, column=5, value=stats['prohibited_count'])
                summary_ws.cell(row=summary_row, column=6, value=stats['normative_count'])
                summary_ws.cell(row=summary_row, column=7, value=stats['ok_count'])
                
                # Apply borders to all cells in the row
                for col in range(1, len(summary_headers) + 1):
                    summary_ws.cell(row=summary_row, column=col).border = summary_border
                
                summary_row += 1
            
            # Set column widths for summary sheet
            summary_column_widths = {
                1: 40,  # Page URL
                2: 20,  # Total Words
                3: 18,  # Unique Words
                4: 20,  # Foreign Words
                5: 20,  # Prohibited Words
                6: 20,  # Normative Violations
                7: 20   # Compliant Words
            }
            for col_idx, width in summary_column_widths.items():
                summary_ws.column_dimensions[get_column_letter(col_idx)].width = width
            
            # Freeze header row in summary sheet
            summary_ws.freeze_panes = "A2"
            
            # Add auto-filter to summary sheet
            if len(page_stats) > 0:
                last_summary_row = summary_row - 1
                summary_ws.auto_filter.ref = f"A1:{get_column_letter(len(summary_headers))}{last_summary_row}"
            
            logger.info(f"ExcelExporter: Summary sheet created with {len(page_stats)} pages")

            logger.info(f"ExcelExporter: About to save workbook to BytesIO buffer")
            # Save to a BytesIO buffer to avoid filesystem corruption
            buffer = io.BytesIO()
            wb.save(buffer)
            logger.info(f"ExcelExporter: Workbook saved to buffer, getting bytes")
            data = buffer.getvalue()
            logger.info(f"ExcelExporter: Got {len(data)} bytes from buffer")
            buffer.close()
            wb.close()
            logger.info(f"ExcelExporter: Returning data to caller")
            return data
            
        except Exception as e:
            logger.error(f"Error generating XLSX export: {e}")
            raise
